#!/usr/bin/env python3
"""
Update SEO Keywords Excel and HTML files with new keywords
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re

BASE_DIR = r"C:\Users\komal\Desktop\07_项目网站\laser-marking-website\(最终)kvlasermarking.com"
OLD_EXCEL = os.path.join(BASE_DIR, "SEO关键词清单.xlsx")
NEW_EXCEL = os.path.join(BASE_DIR, "SEO关键词清单_完整版.xlsx")

# 新增关键词数据
NEW_KEYWORDS = [
    # 应用场景词 - 断桥铝窗厂家
    [57, "首页", "应用场景词", "laser marking for aluminum window profiles", "铝窗型材激光打标", "精准应用搜索 - 窗厂采购", "中"],
    [58, "首页", "应用场景词", "coding for construction aluminum profiles", "建筑铝型材打码", "建筑行业搜索", "中"],
    [59, "首页", "应用场景词", "window frame traceability marking", "窗框批次追溯打标", "质量追溯需求", "低"],
    [60, "首页", "应用场景词", "aluminum door frame laser coder", "铝门框激光打码机", "门框生产搜索", "低"],
    
    # 应用场景词 - 隔热条厂家
    [61, "首页", "应用场景词", "thermal break strip production line coding", "隔热条生产线打码", "产线配套搜索", "中"],
    [62, "首页", "应用场景词", "insulation bar batch production marking", "隔热条批量生产打标", "批量加工搜索", "中"],
    [63, "首页", "应用场景词", "PA66 strip laser marking", "PA66隔热条激光打标", "材质特定搜索", "低"],
    [64, "首页", "应用场景词", "thermal strip reel coding machine", "隔热条盘料打码机", "盘料加工搜索", "低"],
    
    # 技术参数词
    [65, "首页", "技术参数词", "fiber laser marking machine", "光纤激光打标机", "行业通用大词", "高"],
    [66, "首页", "技术参数词", "20W 30W laser coder", "20W/30W激光打码机", "功率参数搜索", "中"],
    [67, "首页", "技术参数词", "permanent non-contact marking", "永久非接触打标", "工艺优势搜索", "低"],
    [68, "首页", "技术参数词", "high speed laser coding system", "高速激光打码系统", "高效率需求", "中"],
    [69, "首页", "技术参数词", "IPG source laser marker", "IPG光源激光打标机", "核心部件搜索", "中"],
    [70, "首页", "技术参数词", "0.01mm precision laser marking", "0.01mm精度激光打标", "高精度需求", "低"],
    
    # 认证词 - 中东/欧美
    [71, "首页", "认证地域词", "CE certified laser marker", "CE认证激光打标机", "准入证搜索", "中"],
    [72, "首页", "认证地域词", "ISO standard laser coding machine", "ISO标准激光打码机", "标准合规搜索", "低"],
    [73, "首页", "认证地域词", "European standard thermal strip coder", "欧洲标准隔热条打码机", "欧标搜索", "低"],
    
    # 地域词 - 东南亚
    [74, "首页", "认证地域词", "thermal strip marking supplier Southeast Asia", "东南亚隔热条打标供应商", "区域供应商搜索", "低"],
    [75, "首页", "认证地域词", "laser coding machine for Malaysia/Thailand/Vietnam", "马来西亚/泰国/越南激光打码机", "东南亚国家搜索", "低"],
    [76, "首页", "认证地域词", "affordable laser marker for Asia market", "亚洲市场高性价比激光打标机", "性价比搜索", "低"],
    
    # 补充高价值词
    [77, "首页", "高价值长尾", "laser marking machine for thermal break strip factory", "断桥铝厂激光打标机", "工厂直购搜索", "低"],
    [78, "首页", "高价值长尾", "insulation strip manufacturer laser coder", "隔热条厂家激光打码机", "厂家配套搜索", "低"],
    [79, "首页", "高价值长尾", "automatic thermal strip coding solution", "隔热条自动打码解决方案", "方案采购搜索", "低"],
    [80, "首页", "高价值长尾", "laser marker for aluminum alloy thermal break", "铝合金断桥隔热条激光打标", "材质+应用搜索", "低"],
]

def update_excel():
    """更新 Excel 文件"""
    print("📊 更新 Excel 表格...")
    
    # 加载现有工作簿
    wb = openpyxl.load_workbook(OLD_EXCEL)
    ws = wb.active
    
    # 样式定义
    data_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 追加数据
    start_row = ws.max_row + 1
    for row_data in NEW_KEYWORDS:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 竞争程度颜色
            if col_idx == 7:
                if value == "高":
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                elif value == "中":
                    cell.fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid")
                elif value == "低":
                    cell.fill = PatternFill(start_color="99CC99", end_color="99CC99", fill_type="solid")
            
        start_row += 1
    
    # 更新自动筛选范围
    ws.auto_filter.ref = f"A1:G{start_row-1}"
    
    wb.save(NEW_EXCEL)
    print(f"✅ Excel 已更新，新增 {len(NEW_KEYWORDS)} 个关键词")
    print(f"📊 总计关键词数: {ws.max_row - 1}")

def update_html_meta():
    """更新 HTML 文件的 meta 标签"""
    print("\n🌐 更新 HTML 文件 meta 标签...")
    
    # 新增关键词按页面分类
    PAGE_KEYWORDS = {
        "index.html": {
            "new_keywords": [
                "laser marking for aluminum window profiles",
                "coding for construction aluminum profiles",
                "thermal break strip production line coding",
                "PA66 strip laser marking",
                "fiber laser marking machine",
                "IPG source laser marker",
                "CE certified laser marker",
                "thermal strip marking supplier Southeast Asia"
            ],
            "new_desc_additions": [
                "CE certified, IPG source fiber laser",
                "for aluminum window manufacturers & thermal strip producers",
                "serving Southeast Asia, Middle East, Europe & North America"
            ]
        },
        "product-1.html": {
            "new_keywords": [
                "PA66 strip laser marking",
                "fiber laser marking machine",
                "permanent non-contact marking"
            ],
            "new_desc_additions": [
                "CE certified fiber laser",
                "ideal for thermal strip manufacturers"
            ]
        },
        "product-2.html": {
            "new_keywords": [
                "thermal break strip production line coding",
                "20W 30W laser coder",
                "high speed laser coding system"
            ],
            "new_desc_additions": [
                "stable production line coding",
                "for batch manufacturing"
            ]
        },
        "product-3.html": {
            "new_keywords": [
                "insulation bar batch production marking",
                "0.01mm precision laser marking",
                "window frame traceability marking"
            ],
            "new_desc_additions": [
                "high precision batch marking",
                "perfect for quality traceability"
            ]
        },
        "product-4.html": {
            "new_keywords": [
                "high speed laser coding system",
                "IPG source laser marker",
                "ISO standard laser coding machine"
            ],
            "new_desc_additions": [
                "flagship high-speed system",
                "ISO standard quality",
                "for large-scale production"
            ]
        }
    }
    
    for filename, data in PAGE_KEYWORDS.items():
        filepath = os.path.join(BASE_DIR, filename)
        if not os.path.exists(filepath):
            print(f"⚠️  文件不存在: {filename}")
            continue
        
        # 尝试不同编码读取
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                html = f.read()
            file_encoding = 'utf-8'
        except UnicodeDecodeError:
            with open(filepath, 'r', encoding='gbk', errors='replace') as f:
                html = f.read()
            file_encoding = 'gbk'
        
        # 1. 更新 keywords
        if '<meta name="keywords"' in html:
            # 提取现有关键词
            match = re.search(r'<meta name="keywords" content="(.*?)"', html)
            if match:
                existing_keywords = match.group(1)
                # 追加新关键词
                new_kw = ", ".join(data["new_keywords"])
                updated_keywords = existing_keywords + ", " + new_kw
                
                # 替换
                html = html.replace(
                    f'<meta name="keywords" content="{existing_keywords}"',
                    f'<meta name="keywords" content="{updated_keywords}"'
                )
                print(f"  ✅ {filename} - 已追加 {len(data['new_keywords'])} 个关键词")
        
        # 2. 更新 description
        if '<meta name="description"' in html:
            match = re.search(r'<meta name="description" content="(.*?)"', html)
            if match:
                existing_desc = match.group(1)
                # 追加描述补充
                new_desc_parts = ". ".join(data["new_desc_additions"])
                updated_desc = existing_desc + ". " + new_desc_parts
                
                html = html.replace(
                    f'<meta name="description" content="{existing_desc}"',
                    f'<meta name="description" content="{updated_desc}"'
                )
                print(f"  ✅ {filename} - 已更新 description")
        
        with open(filepath, 'w', encoding=file_encoding) as f:
            f.write(html)
    
    print(f"\n✅ 共更新 {len(PAGE_KEYWORDS)} 个 HTML 文件")

if __name__ == "__main__":
    print("🚀 开始更新 SEO关键词...")
    print("="*50)
    
    # 1. 更新 Excel
    update_excel()
    
    # 2. 更新 HTML
    update_html_meta()
    
    print("\n" + "="*50)
    print("✨ 所有更新完成！")
    print(f"📊 关键词总数: {80} 个")
    print(f"🌐 HTML 文件已同步更新")
