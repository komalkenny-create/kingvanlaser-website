#!/usr/bin/env python3
"""
Generate SEO Keywords Excel for kvlasermarking.com
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SEO关键词清单"

# 样式定义
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
category_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
category_font = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
data_font = Font(name="微软雅黑", size=10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表头
headers = ["序号", "页面", "关键词类型", "英文关键词", "中文翻译", "搜索意图", "竞争程度"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 关键词数据
keywords_data = [
    # 首页核心词
    [1, "首页", "核心大词", "laser marking machine for thermal strips", "隔热条激光打码机", "行业核心词 - 设备采购", "高"],
    [2, "首页", "核心大词", "thermal break strip coding", "隔热条打码", "精准业务词 - 技术搜索", "中"],
    [3, "首页", "核心大词", "insulation bar laser marking", "隔热条激光标记", "精准业务词 - 材料搜索", "中"],
    [4, "首页", "核心大词", "automatic laser coder", "自动激光打码机", "设备采购词 - 自动化需求", "中"],
    [5, "首页", "核心大词", "aluminum profile marking", "铝型材标记", "应用场景词 - 行业搜索", "中"],
    [6, "首页", "品牌词", "KINGVAN laser", "KINGVAN 激光", "品牌搜索 - 已知品牌", "低"],
    [7, "首页", "品牌词", "KINGVAN Technology", "KINGVAN 科技", "品牌搜索 - 公司名", "低"],
    [8, "首页", "长尾词", "laser marking solution for insulation strips", "隔热条激光标记解决方案", "解决方案搜索", "低"],
    [9, "首页", "长尾词", "automatic thermal strip coding machine", "自动隔热条打码机", "精准设备搜索", "低"],
    [10, "首页", "长尾词", "laser coder for aluminum window profiles", "铝窗型材激光打码机", "应用场景搜索", "低"],
    
    # Product 1 - 单面单牵引
    [11, "Product 1", "核心词", "single side laser marking", "单面激光打标", "单面打标需求", "中"],
    [12, "Product 1", "核心词", "single traction laser coder", "单牵引激光打码机", "单牵引设备搜索", "中"],
    [13, "Product 1", "长尾词", "economical laser coder", "经济型激光打码机", "预算敏感型采购", "低"],
    [14, "Product 1", "长尾词", "single side single traction laser marking", "单面单牵引激光打标", "精准型号搜索", "低"],
    [15, "Product 1", "长尾词", "basic thermal strip marking machine", "基础型隔热条打标机", "入门级设备搜索", "低"],
    [16, "Product 1", "长尾词", "affordable laser coder for thermal strips", "经济型隔热条激光打码机", "价格敏感搜索", "低"],
    
    # Product 2 - 单面双牵引
    [17, "Product 2", "核心词", "single side double traction", "单面双牵引", "双牵引设备搜索", "中"],
    [18, "Product 2", "核心词", "dual traction laser marking", "双牵引激光打标", "稳定送料需求", "中"],
    [19, "Product 2", "长尾词", "long strip laser coder", "长条激光打码机", "长料加工搜索", "低"],
    [20, "Product 2", "长尾词", "stable feeding laser machine", "稳定送料激光机", "稳定性需求", "低"],
    [21, "Product 2", "长尾词", "double traction thermal strip coder", "双牵引隔热条打码机", "精准型号搜索", "低"],
    [22, "Product 2", "长尾词", "enhanced stability laser marking", "增强稳定性激光打标", "高性能需求", "低"],
    
    # Product 3 - 双面单牵引
    [23, "Product 3", "核心词", "double side laser marking", "双面激光打标", "双面打标需求", "中"],
    [24, "Product 3", "核心词", "dual side laser coder", "双面激光打码机", "双面设备搜索", "中"],
    [25, "Product 3", "长尾词", "simultaneous marking machine", "同步打标机", "高效率需求", "低"],
    [26, "Product 3", "长尾词", "thermal strip double coding", "隔热条双面打码", "双面工艺搜索", "低"],
    [27, "Product 3", "长尾词", "insulation bar dual marking", "隔热条双面标记", "双面加工搜索", "低"],
    [28, "Product 3", "长尾词", "both side laser marking machine", "双面激光打标机", "精准需求搜索", "低"],
    
    # Product 4 - 双面双牵引（旗舰）
    [29, "Product 4", "核心词", "double side double traction", "双面双牵引", "高端设备搜索", "中"],
    [30, "Product 4", "核心词", "flagship laser marking", "旗舰激光打标", "高端采购搜索", "低"],
    [31, "Product 4", "长尾词", "high-speed laser coder", "高速激光打码机", "高效率需求", "中"],
    [32, "Product 4", "长尾词", "automatic thermal strip machine", "自动隔热条设备", "全自动需求", "低"],
    [33, "Product 4", "长尾词", "premium laser marking system", "高端激光打标系统", "高端采购搜索", "低"],
    [34, "Product 4", "长尾词", "high-volume thermal strip production", "大批量隔热条生产", "产线采购搜索", "低"],
    
    # About 页面
    [35, "About", "品牌词", "KINGVAN Technology company", "KINGVAN 科技公司", "公司背景搜索", "低"],
    [36, "About", "行业词", "laser marking manufacturer", "激光打标机制造商", "厂家搜索", "中"],
    [37, "About", "行业词", "thermal strip coding company", "隔热条打码公司", "供应商搜索", "低"],
    [38, "About", "品牌词", "about KINGVAN", "关于 KINGVAN", "品牌了解搜索", "低"],
    [39, "About", "长尾词", "KINGVAN laser marking factory", "KINGVAN 激光打标工厂", "工厂直搜", "低"],
    
    # Blog 页面
    [40, "Blog", "内容词", "laser marking blog", "激光打标博客", "技术文章搜索", "中"],
    [41, "Blog", "内容词", "thermal strip coding guide", "隔热条打码指南", "教程搜索", "低"],
    [42, "Blog", "内容词", "laser technology news", "激光技术新闻", "行业动态搜索", "中"],
    [43, "Blog", "内容词", "insulation strip tips", "隔热条技巧", "经验分享搜索", "低"],
    [44, "Blog", "内容词", "KINGVAN articles", "KINGVAN 文章", "品牌内容搜索", "低"],
    
    # News 页面
    [45, "News", "内容词", "KINGVAN news", "KINGVAN 新闻", "公司动态搜索", "低"],
    [46, "News", "内容词", "laser marking industry news", "激光打标行业新闻", "行业新闻搜索", "中"],
    [47, "News", "内容词", "thermal strip updates", "隔热条更新", "产品动态搜索", "低"],
    [48, "News", "内容词", "company announcements", "公司公告", "官方信息发布", "低"],
    
    # 额外高价值长尾词
    [49, "首页", "高价值长尾", "laser marking machine price thermal strip", "隔热条激光打码机价格", "价格搜索 - 高转化", "中"],
    [50, "首页", "高价值长尾", "best laser coder for insulation strips", "最佳隔热条激光打码机", "对比搜索 - 高转化", "低"],
    [51, "首页", "高价值长尾", "thermal break strip laser marking supplier", "隔热条激光打标供应商", "供应商搜索", "低"],
    [52, "首页", "高价值长尾", "automatic laser marking system for thermal strips", "隔热条自动激光打标系统", "系统采购搜索", "低"],
    [53, "首页", "高价值长尾", "laser coding machine for aluminum window insulation", "铝窗隔热激光打码机", "精准应用搜索", "低"],
    [54, "首页", "高价值长尾", "thermal strip laser marking China manufacturer", "隔热条激光打标中国制造商", "源头工厂搜索", "低"],
    [55, "首页", "高价值长尾", "laser marker for PVC thermal break strips", "PVC隔热条激光标记机", "材料特定搜索", "低"],
    [56, "首页", "高价值长尾", "continuous laser marking for insulation bars", "隔热条连续激光打标", "工艺特定搜索", "低"],
]

# 填充数据
for row_idx, row_data in enumerate(keywords_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 根据竞争程度设置颜色
        if col_idx == 7:  # 竞争程度列
            if value == "高":
                cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            elif value == "中":
                cell.fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid")
            elif value == "低":
                cell.fill = PatternFill(start_color="99CC99", end_color="99CC99", fill_type="solid")

# 设置列宽
ws.column_dimensions['A'].width = 6   # 序号
ws.column_dimensions['B'].width = 15  # 页面
ws.column_dimensions['C'].width = 15  # 关键词类型
ws.column_dimensions['D'].width = 45  # 英文关键词
ws.column_dimensions['E'].width = 30  # 中文翻译
ws.column_dimensions['F'].width = 30  # 搜索意图
ws.column_dimensions['G'].width = 12  # 竞争程度

# 设置行高
for row in ws.iter_rows(min_row=1, max_row=len(keywords_data)+1):
    ws.row_dimensions[row[0].row].height = 25

# 冻结首行
ws.freeze_panes = 'A2'

# 自动筛选
ws.auto_filter.ref = f"A1:G{len(keywords_data)+1}"

# 保存文件
output_path = r"C:\Users\komal\Desktop\07_项目网站\laser-marking-website\(最终)kvlasermarking.com\SEO关键词清单.xlsx"
wb.save(output_path)
print(f"✅ Excel 文件已生成: {output_path}")
print(f"📊 共 {len(keywords_data)} 个关键词")
